"""services/export.py — chiqadigan fayl namuna POVTOR fayli bilan bir xil bo'lishi."""

from __future__ import annotations

import io
from datetime import date

REPORT_DATE = date.today()

import pytest
from openpyxl import load_workbook

from povtor_bot.core.models import STATUS_NOT_FOUND, STATUS_TAKEN
from povtor_bot.db import repo
from povtor_bot.services import export

from .conftest import make_candidate

pytestmark = pytest.mark.usefixtures("database")


async def _seed_and_answer(pairs: list[tuple[str, str, str]]) -> list:
    """pairs: (sku, filial nomi, status).

    Export JAVOB BERILGAN kun bo'yicha ishlaydi, shuning uchun hisobot sanasi —
    haqiqiy bugun (javob shu daqiqada yoziladi), test fixture'idagi TODAY emas.
    """
    await repo.insert_candidates([
        make_candidate(sku=sku, shop_id=shop.lower(), shop_name=shop)
        for sku, shop, _ in pairs
    ])
    for sku, shop, status in pairs:
        rows = [r for r in await repo.card_items(sku) if r["shop_name"] == shop]
        await repo.answer_candidate(rows[0]["id"], status=status, user_id=1)
    return await repo.answered_for_export(REPORT_DATE)


def _load(content: bytes):
    return load_workbook(io.BytesIO(content))


class TestWorkbookStructure:
    async def test_one_sheet_per_filial(self) -> None:
        rows = await _seed_and_answer([
            ("1", "ANDALUS", STATUS_TAKEN),
            ("2", "BERUNIY", STATUS_TAKEN),
            ("3", "ANDALUS", STATUS_NOT_FOUND),
        ])
        book = _load(export.build_workbook(rows, REPORT_DATE))
        assert set(book.sheetnames) == {"ANDALUS", "BERUNIY"}

    async def test_header_row_matches_sample_file(self) -> None:
        """14 ustun, sarlavha 3-qatorda — namuna fayldagidek."""
        rows = await _seed_and_answer([("1", "ANDALUS", STATUS_TAKEN)])
        sheet = _load(export.build_workbook(rows, REPORT_DATE))["ANDALUS"]
        headers = [sheet.cell(3, i).value for i in range(1, 15)]
        assert headers == export.HEADERS
        assert len(export.HEADERS) == 14

    async def test_title_row_carries_date_and_count(self) -> None:
        rows = await _seed_and_answer([
            ("1", "ANDALUS", STATUS_TAKEN), ("2", "ANDALUS", STATUS_TAKEN),
        ])
        sheet = _load(export.build_workbook(rows, REPORT_DATE))["ANDALUS"]
        title = sheet.cell(1, 1).value
        assert "ANDALUS" in title and REPORT_DATE.isoformat() in title and "2 poz" in title


class TestAnswerColumns:
    async def test_taken_fills_quantity_column(self) -> None:
        rows = await _seed_and_answer([("1", "ANDALUS", STATUS_TAKEN)])
        sheet = _load(export.build_workbook(rows, REPORT_DATE))["ANDALUS"]
        assert sheet.cell(4, 13).value == 10        # OLINDI (dona) = tavsiya
        assert sheet.cell(4, 14).value is None      # BOZORDA YO'Q bo'sh

    async def test_not_found_marks_cross(self) -> None:
        rows = await _seed_and_answer([("1", "ANDALUS", STATUS_NOT_FOUND)])
        sheet = _load(export.build_workbook(rows, REPORT_DATE))["ANDALUS"]
        assert sheet.cell(4, 13).value is None
        assert sheet.cell(4, 14).value == "×"

    async def test_data_columns_carry_statistics(self) -> None:
        rows = await _seed_and_answer([("39666", "ANDALUS", STATUS_TAKEN)])
        sheet = _load(export.build_workbook(rows, REPORT_DATE))["ANDALUS"]
        values = [sheet.cell(4, i).value for i in range(1, 13)]
        assert values[0] == "39666"                 # Artikul
        assert values[1] == "Белый"                 # Rang
        assert values[2] == "Рубашка с дл/р"        # Podkategoriya
        assert values[4] == "Sharof M255"           # Postavshik
        assert values[5:9] == [5, 5, 100.0, 2]      # Asos, Sotilgan, Foiz, kun
        assert values[9] == "ishonchli"
        assert values[10] == 10                     # Bugun beriladi


class TestSheetNames:
    def test_truncates_and_dedupes(self) -> None:
        used: set[str] = set()
        long_name = "A" * 40
        first = export.safe_sheet_name(long_name, used)
        second = export.safe_sheet_name(long_name, used)
        assert len(first) <= 31 and len(second) <= 31
        assert first != second

    def test_strips_invalid_characters(self) -> None:
        assert export.safe_sheet_name("A/B:C[1]", set()) == "A-B-C-1-"

    def test_empty_name_gets_placeholder(self) -> None:
        assert export.safe_sheet_name("   ", set()) == "Filial"


class TestEdgeCases:
    def test_empty_export_still_valid_file(self) -> None:
        book = _load(export.build_workbook([], date(2026, 8, 20)))
        assert book.sheetnames == ["Bo'sh"]

    def test_filename_format(self) -> None:
        assert export.filename(date(2026, 8, 20)) == "POVTOR_2026-08-20.xlsx"
