"""Kunning javoblarini Excel'ga chiqarish.

Ustunlar va varaqlar tuzilishi hozirgi qo'lda ishlatilayotgan POVTOR fayli
bilan AYNAN bir xil — menejerlar formatni qayta o'rganmasin, va eski
fayllar bilan solishtirish mumkin bo'lsin.
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.models import STATUS_TAKEN

HEADERS = [
    "Artikul", "Rang", "Podkategoriya", "Tur", "Postavshik", "Asos", "Sotilgan",
    "Foiz", "50% ga yetgan kun", "Daraja", "Bugun beriladi (dona)", "Izoh",
    "OLINDI (dona)", "BOZORDA YO'Q (×)",
]
# Namuna faylda sarlavha 3-qatorda, 1-qator — izoh matni
_TITLE_ROW = 1
_HEADER_ROW = 3

_COLUMN_WIDTHS = [10, 16, 22, 14, 22, 8, 10, 8, 18, 12, 20, 30, 14, 16]
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
# Excel varaq nomida taqiqlangan belgilar
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel varaq nomi: 31 belgigacha, taqiqlangan belgilarsiz, takrorlanmas."""
    cleaned = _INVALID_SHEET_CHARS.sub("-", (name or "").strip()) or "Filial"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate.casefold() in used:
        tail = f"_{suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def build_workbook(rows: Sequence[Any], report_date: date) -> bytes:
    """Har bir filial uchun alohida varaq. Qatorsiz filial varaq ochmaydi."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    by_shop: dict[str, list[Any]] = {}
    for row in rows:
        by_shop.setdefault(row["shop_name"], []).append(row)

    used: set[str] = set()
    for shop_name, shop_rows in sorted(by_shop.items()):
        sheet = workbook.create_sheet(safe_sheet_name(shop_name, used))
        sheet.cell(_TITLE_ROW, 1).value = (
            f"POVTOR — {shop_name} · {report_date.isoformat()} · {len(shop_rows)} poz"
        )
        sheet.cell(_TITLE_ROW, 1).font = Font(bold=True, size=12)

        for index, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(_HEADER_ROW, index)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for offset, row in enumerate(shop_rows, start=1):
            taken = row["status"] == STATUS_TAKEN
            values = [
                row["sku"], row["color"], row["subcategory"], row["kind"],
                row["supplier"],
                row["base_qty"], row["sold_qty"], row["percent"], row["days_to_50"],
                row["grade"], row["recommended_qty"], row["note"],
                # Bo'sh katak None bilan yoziladi: openpyxl "" ni baribir bo'sh
                # katak qilib o'qiydi, None esa niyatni aniq ko'rsatadi
                row["recommended_qty"] if taken else None,
                None if taken else "×",
            ]
            for index, value in enumerate(values, start=1):
                sheet.cell(_HEADER_ROW + offset, index).value = value

        for index, width in enumerate(_COLUMN_WIDTHS, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = sheet.cell(_HEADER_ROW + 1, 1)

    if not workbook.sheetnames:
        # Bo'sh fayl ham yaroqli bo'lishi kerak — openpyxl varaqsiz saqlamaydi
        sheet = workbook.create_sheet("Bo'sh")
        sheet.cell(1, 1).value = f"{report_date.isoformat()} uchun javob berilgan band yo'q"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def filename(report_date: date) -> str:
    return f"POVTOR_{report_date.isoformat()}.xlsx"
